import pandas as pd
import numpy as np
import openpyxl
from sklearn.neighbors import BallTree
from sklearn.base import BaseEstimator
from sklearn.pipeline import make_pipeline
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.decomposition import TruncatedSVD
from aiogram import Bot, Dispatcher, executor, types
from aiogram.dispatcher.filters import BoundFilter
import logging
import random


df = pd.read_excel('dataframe.xlsx')
wb = openpyxl.load_workbook('dataframe.xlsx')
my_sheet = wb.active

vektor = TfidfVectorizer()
vektor.fit(df.context_0)
matrix_big = vektor.transform(df.context_0)

svd = TruncatedSVD()
svd.fit(matrix_big)
matrix_small = svd.transform(matrix_big)

def soft_max(x):
    proba = np.exp(-x)
    return proba/sum(proba)

class NeighborSampler(BaseEstimator):
    def __init__(self,k=1,temperature=1):
        self.k = k
        self.temperature = temperature
    def fit(self,X,y):
        self.tree_ = BallTree(X)
        self.y_ = np.array(y)
    def predict(self, X, random_state = None):
        distances, indices = self.tree_.query(X, return_distance=True,k=self.k)
        result = []
        for distance, index in zip(distances, indices):
            result.append(np.random.choice(index, p=soft_max(distance*self.temperature)))
            return self.y_[result]

ns = NeighborSampler()
ns.fit(matrix_small, df.reply)
pipe = make_pipeline(vektor, svd, ns)


logging.basicConfig(level=logging.INFO)

bot = Bot(token="token")
dp = Dispatcher(bot)


group_id = "enter the group_id without quotation marks"

class IsAdminFilter(BoundFilter):
    key = "is_admin"

    def __init__(self,is_admin):
        self.is_admin = is_admin

    async def check(self, message: types.Message):
        member = await message.bot.get_chat_member(message.chat.id,message.from_user.id)
        return member.is_chat_admin()

dp.filters_factory.bind(IsAdminFilter)


@dp.message_handler(commands=['start'])
async def send_welcome(message: types.Message):
    user_idd = message.from_user.first_name
    await message.answer(f'Bot Adolf welcomes {user_idd}!')



@dp.message_handler(is_admin=True, commands=['ban', 'kick'])
async def cmd_ban(message):
    if not message.reply_to_message:
        await message.reply("reply to the message of the person who must be banned")
        return
    await message.bot.delete_message(group_id, message.message_id)
    await message.bot.kick_chat_member(group_id, user_id=message.reply_to_message.from_user.id)

    await message.reply_to_message.reply(f"{message.reply_to_message.from_user.first_name} was successful banned ")

messagess = 0
voicemessages = 0
photos = 0
videos = 0

@dp.message_handler(commands=["statistic"])
async def statistics(message: types.Message):
    await message.answer(f"In this chat you wrote {messagess} messages, {photos} photos, {voicemessages} voice messages and {videos} videos")

@dp.message_handler(commands=["when"], commands_prefix='.')
async def count_date(message: types.Message):
    year = random.randint(2024,3000)
    month = random.randint(1,12)
    day = random.randint(1,30)
    a = message.text
    a.split()
    await message.reply(f"{day}.{month}.{year} ")


#photos counter
@dp.message_handler(content_types=["photo"])
async def count_photos(message):
    global photos
    photos = photos + 1

#voice messages counter
@dp.message_handler(content_types=["voice"])
async def count_voice(message):
    global voicemessages
    voicemessages = voicemessages + 1


#videos counter
@dp.message_handler(content_types=["video"])
async def count_videos(message):
    global videos
    videos = videos + 1

enabled = False

#AI enabling
@dp.message_handler(commands=["hi"])
async def enable_AI(message: types.Message):
    global enabled
    if enabled is False:
        enabled = True
        await message.answer("Bot Adolf is on")
    elif enabled is True:
        enabled = False
        await message.answer("Bot Adolf is turned off")


#messages counter and entering messages into the db
@dp.message_handler(content_types=["text"])
async def valueEditor(message: types.Message):
    global messagess
    messagess = messagess + 1
    if enabled == 0:
        row = len(list(my_sheet.rows)) + 1
        c1 = my_sheet.cell(row=row, column=3)
        c1.value = message.text
        c2 = my_sheet.cell(row=row, column=4)
        c2.value = message.text
        wb.save("./dataframe.xlsx")
    else:
        await message.answer(pipe.predict([message.text.lower()])[0])

if __name__ == "__main__":
    executor.start_polling(dp, skip_updates=True)

